#!/usr/bin/env python
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from collections import namedtuple
from dataclasses import dataclass
from math import ceil
import sys


@dataclass(frozen=True)
class Actions:
    Marker: str = "Marker"
    Image: str = "Image"
    Video: str = "Video"


@dataclass(frozen=True)
class Colors:
    Blue: str = "Blue"
    Cyan: str = "Cyan"
    Green: str = "Green"
    Yellow: str = "Yellow"
    Red: str = "Red"
    Pink: str = "Pink"
    Purple: str = "Purple"
    Fuchsia: str = "Fuchsia"
    Rose: str = "Rose"
    Lavender: str = "Lavender"
    Sky: str = "Sky"
    Mint: str = "Mint"
    Lemon: str = "Lemon"
    Sand: str = "Sand"
    Cocoa: str = "Cocoa"
    Cream: str = "Cream"


@dataclass(frozen=True)
class Media:
    index: int
    startFrame: int
    insertFrame: int
    duration: int
    action: str
    color: str
    name: str
    note: str
    path: str


def GetResolve():
    try:
        # The PYTHONPATH needs to be set correctly for this import statement to work.
        # An alternative is to import the DaVinciResolveScript by specifying absolute path (see ExceptionHandler logic)
        import DaVinciResolveScript as bmd
    except ImportError:
        if sys.platform.startswith("darwin"):
            expectedPath = "/Library/Application Support/Blackmagic Design/DaVinci Resolve/Developer/Scripting/Modules/"
        elif sys.platform.startswith("win") or sys.platform.startswith("cygwin"):
            import os

            expectedPath = (
                os.getenv("PROGRAMDATA")
                + "\\Blackmagic Design\\DaVinci Resolve\\Support\\Developer\\Scripting\\Modules\\"
            )
        elif sys.platform.startswith("linux"):
            expectedPath = "/opt/resolve/libs/Fusion/Modules/"

        # check if the default path has it...
        print(
            "Unable to find module DaVinciResolveScript from $PYTHONPATH - trying default locations"
        )
        try:
            import imp

            bmd = imp.load_source(
                "DaVinciResolveScript", expectedPath + "DaVinciResolveScript.py"
            )
        except ImportError:
            # No fallbacks ... report error:
            print(
                "Unable to find module DaVinciResolveScript - please ensure that the module DaVinciResolveScript is discoverable by python"
            )
            print(
                "For a default DaVinci Resolve installation, the module is expected to be located in: "
                + expectedPath
            )
            sys.exit()

    return bmd.scriptapp("Resolve")


def CreateConnection():
    try:
        resolve = GetResolve()
        projectManager = resolve.GetProjectManager()
        return projectManager.GetCurrentProject()
    except:
        messagebox.showerror(
            "Connection error", "Couldn't connect to Resolve, check API installation"
        )
        sys.exit()


def DisplayProjectInfo(project):
    print("-----------")
    print("Project '" + project.GetName() + "':")
    framerate = ceil(project.GetSetting("timelineFrameRate"))
    print("  Framerate " + str(framerate))
    print(
        "  Resolution "
        + project.GetSetting("timelineResolutionWidth")
        + "x"
        + project.GetSetting("timelineResolutionHeight")
    )
    return framerate


def GetTimeline(project):
    timeline = project.GetCurrentTimeline()
    if timeline != None:
        print(f"Active timeline name: {timeline.GetName()}")
        return timeline
    else:
        messagebox.showerror("Timeline error", "Timeline is not selected/created.")
        sys.exit()


def GetExcelWorksheet():
    try:
        path = filedialog.askopenfilename()
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        print(f"Worksheet '{ws.title}' loaded.")
        return ws
    except:
        messagebox.showerror("Excel error", "Couldn't load Excel worksheet.")
        sys.exit()


def PlaceMarker(timeline, media: Media):

    # timeline.SetCurrentTimecode("01:00:02:16")
    # tli = timeline.GetCurrentVideoItem()
    # tli.AddMarker(64, "Blue", "Name3", "Note3", 1)

    # m_pos = 64
    # tl_name = timeline.GetName()
    # tl_start = timeline.GetStartFrame()
    # timeline.SetCurrentTimecode("01:00:00:00") # 64
    # tli = timeline.GetCurrentVideoItem()
    # tli_start = tli.GetStart()

    # print(tli_start - tl_start)
    # print(tli.GetLeftOffset())
    # pos = tli.GetLeftOffset() - (tli_start - tl_start) + m_pos
    # print(pos)

    # tli.AddMarker(pos, "Blue", "Name3", "Note3", 1)

    timeline.AddMarker(
        media.insertFrame, media.color, media.name, media.note, media.duration
    )
    print(f"{media.index}) Added marker '{media.name}' at frame {media.insertFrame}")


def ProcessWorksheet(worksheet, timeline):
    for raw_row in worksheet.iter_rows(min_row=2):  # skip header
        media = Media(*[cell.value for cell in raw_row])
        if media.action == actions.Marker:
            PlaceMarker(timeline, media)
        else:
            print("No action found, exiting...")
            break


if __name__ == "__main__":
    actions = Actions()

    project = CreateConnection()
    framerate = DisplayProjectInfo(project)
    timeline = GetTimeline(project)

    # tl_item = timeline.GetItemListInTrack("video", 1)
    # for item in tl_item:
    #     print(item.GetName())
    #     print(item.GetStart())

    worksheet = GetExcelWorksheet()

    ProcessWorksheet(worksheet, timeline)
